import openpyxl  # Import openpyxl to work with Excel files (.xlsx)
from openpyxl.styles import PatternFill  # Import PatternFill to apply cell background colors

def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)  # Load the Excel workbook from the given file path
    sheet = workbook[sheet_name]  # Select the specified sheet by name
    return (sheet.max_row)  # Return the total number of rows in the sheet

def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)  # Load the Excel workbook from the given file path
    sheet = workbook[sheet_name]  # Select the specified sheet by name
    return (sheet.max_column)  # Return the total number of columns in the sheet

def read_data(file, sheet_name, row, column):
    workbook = openpyxl.load_workbook(file)  # Load the Excel workbook from the given file path
    sheet = workbook[sheet_name]  # Select the specified sheet by name
    return sheet.cell(row=row, column=column).value  # Return the value from the specified cell

def write_data(file, sheet_name, row, column, data):
    workbook = openpyxl.load_workbook(file)  # Load the Excel workbook from the given file path
    sheet = workbook[sheet_name]  # Select the specified sheet by name
    sheet.cell(row=row, column=column).value = data  # Write the provided data to the specified cell
    workbook.save(file)  # Save the workbook to persist changes

# def fill_green_color(file, sheet_name, row, column):
#     workbook = openpyxl.load_workbook(file)  # Load the Excel workbook from the given file path
#     sheet = workbook[sheet_name]  # Select the specified sheet by name
#     fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Create a green fill pattern
#     sheet.cell(row=row, column=column).fill = fill  # Apply the green fill to the specified cell
#     workbook.save(file)  # Save the workbook to persist changes

# def fill_red_color(file, sheet_name, row, column):
#     workbook = openpyxl.load_workbook(file)  # Load the Excel workbook from the given file path
#     sheet = workbook[sheet_name]  # Select the specified sheet by name
#     fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Create a red fill pattern
#     sheet.cell(row=row, column=column).fill = fill  # Apply the red fill to the specified cell
#     workbook.save(file)  # Save the workbook to persist